#!/usr/bin/env python3
"""Build a Sales Mapping dataset from a territory spreadsheet.

Reads a sheet with one row per territory — current area, territory name, twelve
monthly revenue columns, then a full-year column — and joins it to the reference
tables in tools/reference/ (coordinates, states covered, adjacency, land area,
planned adds) to produce the JSON the app consumes.

    python3 tools/build_dataset.py book.xlsx out.json \
        --sheet "Territory Summary" --first-row 3 --name "2027 plan"

The starting area assignment (`baseline`) comes from the spreadsheet's own area
column, so you begin from today's structure and re-cut it in the browser.

Anything the reference tables do not know about is reported and the build stops
— a territory with no coordinates cannot be drawn, and one with no adjacency
would silently break the contiguity check.
"""
import argparse
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "reference"))

from coords import LL          # noqa: E402  territory -> (lat, lon)
from states import TS          # noqa: E402  territory -> [state codes]
from geo import ADJ            # noqa: E402  territory -> [adjacent territories]
from land import AREA          # noqa: E402  territory -> land sq mi
from expand import ADDS        # noqa: E402  planned headcount adds

# Six colour-blind-validated pairs, then six fallbacks for areas you create later.
PAL_LIGHT = ['#e19b11', '#b271a8', '#12c6a5', '#c33e0a', '#167d59', '#5060b5',
             '#8a6d3b', '#2f7fb5', '#a8324a', '#5b8c2a', '#7a5195', '#4a4a48']
PAL_DARK = ['#c18408', '#a53d99', '#10ad8f', '#b83804', '#097552', '#4654d8',
            '#a9854a', '#3d9ad6', '#c4405b', '#6fa834', '#9563b4', '#6a6a66']
VALIDATED = 7


def read_sheet(path, sheet, first_row):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(min_row=first_row, max_col=15):
        area, name = r[0].value, r[1].value
        if not name or not area:          # blank row or a total line
            continue
        fy = r[14].value
        if fy is None:
            fy = sum(float(c.value or 0) for c in r[2:14])
        rows.append((str(area).strip(), str(name).strip(), float(fy or 0)))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("out")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--first-row", type=int, default=3)
    ap.add_argument("--name", default="Imported plan")
    a = ap.parse_args()

    rows = read_sheet(a.xlsx, a.sheet, a.first_row)
    names = [n for _, n, _ in rows]

    missing = {
        "coordinates (tools/reference/coords.py)": [n for n in names if n not in LL],
        "states (tools/reference/states.py)": [n for n in names if n not in TS],
        "adjacency (tools/reference/geo.py)": [n for n in names if n not in ADJ],
    }
    bad = {k: v for k, v in missing.items() if v}
    if bad:
        for k, v in bad.items():
            print(f"! no {k} for: {', '.join(v)}", file=sys.stderr)
        sys.exit("add them to the reference tables, then re-run")

    orphan_adds = [x[0] for x in ADDS if x[3] not in names]
    if orphan_adds:
        print(f"! planned adds with no parent territory: {', '.join(orphan_adds)}", file=sys.stderr)
        sys.exit("fix tools/reference/expand.py, then re-run")

    areas = sorted({a_ for a_, _, _ in rows}, key=lambda x: [a_ for a_, _, _ in rows].index(x))
    aidx = {a_: i for i, a_ in enumerate(areas)}
    ti = {n: i for i, n in enumerate(names)}

    territories = []
    for area, name, fy in rows:
        kids = [dict(market=m, role=role, timing=t, lat=la, lon=lo)
                for m, role, t, parent, st, la, lo, why in ADDS if parent == name]
        territories.append(dict(
            name=name, lat=LL[name][0], lon=LL[name][1], rev=round(fy),
            land=round(AREA[name]), states=TS[name], old=area, area=aidx[area],
            heads=1 + len(kids), adds=kids, nAdds=len(kids)))

    flat_adds = [dict(market=k["market"], role=k["role"], timing=k["timing"],
                      lat=k["lat"], lon=k["lon"], parent=i)
                 for i, t in enumerate(territories) for k in t["adds"]]

    all_states = sorted({s for t in territories for s in t["states"]})
    ds = {
        "meta": {"name": a.name, "source": os.path.basename(a.xlsx),
                 "generated_by": "tools/build_dataset.py", "currency": "USD"},
        "territories": territories,
        "areas": [dict(name=a_, light=PAL_LIGHT[i % 12], dark=PAL_DARK[i % 12],
                       leader="", role="AVP", open=True,
                       lat=sum(t["lat"] for t in territories if t["area"] == i) /
                           max(1, sum(1 for t in territories if t["area"] == i)),
                       lon=sum(t["lon"] for t in territories if t["area"] == i) /
                           max(1, sum(1 for t in territories if t["area"] == i)),
                       core=[])
                  for i, a_ in enumerate(areas)],
        "adj": {ti[k]: [ti[x] for x in v if x in ti] for k, v in ADJ.items() if k in ti},
        "colocate": [],
        "stateTerr": {s: [i for i, t in enumerate(territories) if s in t["states"]]
                      for s in all_states},
        "uncovered": [s for s in ("AK", "HI") if s not in all_states],
        "adds": flat_adds,
        "total": sum(t["rev"] for t in territories),
        "totalHeads": len(territories) + len(flat_adds),
        "k": len(areas),
        "baseline": [t["area"] for t in territories],
        "palette": {"light": PAL_LIGHT, "dark": PAL_DARK, "validated": VALIDATED},
        "homes": [],
    }
    json.dump(ds, open(a.out, "w"), indent=1)
    print(f"{a.out}: {len(territories)} territories in {len(areas)} areas, "
          f"{len(flat_adds)} planned adds, ${ds['total']:,}")


if __name__ == "__main__":
    main()
